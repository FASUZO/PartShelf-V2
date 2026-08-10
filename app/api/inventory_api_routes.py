"""
库存管理 API 路由模块
提供库存相关的所有API接口，包括：
- 零件添加、查询、删除
- 库存更新（入库/出库）
- 批量导入导出
- 库存历史记录
"""

import io
import logging
from typing import List
from openpyxl import Workbook
from fastapi import APIRouter, Depends, File, Form, Query, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from app.schemas.file_template import FileTemplateAdd
from app.schemas.inventory import PartToInventoryAdd, PartInventoryQuantityUpdate, PartInventoryFilter
from app.models.config import Category, Subcategory
from app.services.file_service import FileService
from app.services.inventory_service import InventoryService
from app.api.deps import get_current_user_required
from db.database import get_db

logger = logging.getLogger("partshelf.api")
router = APIRouter()

# ==================== 零件管理接口 ====================

@router.post("/add_part_to_inventory")
def add_part_to_inventory(
    name: str = Form(...),
    manufacturer: str = Form(...),
    part_type: str = Form(...),
    package: str = Form(...),
    quantity: int = Form(...),
    description: str = Form(None),
    price: float = Form(None),
    lc_number: str = Form(None),
    other: str = Form(None),
    category_id: int = Form(None),
    subcategory_id: int = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user_required)
):
    """添加新零件到库存"""
    part_data = PartToInventoryAdd(
        name=name,
        manufacturer=manufacturer,
        part_type=part_type,
        package=package,
        quantity=quantity,
        description=description,
        price=price,
        lc_number=lc_number,
        other=other,
        category_id=category_id,
        subcategory_id=subcategory_id
    )
    InventoryService.add_part_to_inventory(db, part_data)
    logger.info("零件入库: %s (%s/%s), 数量=%d", name, manufacturer, package, quantity)
    return RedirectResponse("/inventory", status_code=303)

@router.post("/update_quantity")
def update_quantity(
    part_id: int = Form(...),
    quantity_change: int = Form(...),
    remark: str = Form(""),
    db: Session = Depends(get_db),
    user=Depends(get_current_user_required)
):
    """
    更新库存数量
    - quantity_change > 0: 入库
    - quantity_change < 0: 出库
    - quantity_change = 0: 调整
    """
    # 根据数量变化确定操作模式
    if quantity_change > 0:
        operation_mode, quantity = "add", quantity_change
    elif quantity_change < 0:
        operation_mode, quantity = "subtract", abs(quantity_change)
    else:
        operation_mode, quantity = "set", 0
    
    update_data = PartInventoryQuantityUpdate(
        part_id=part_id,
        quantity=quantity,
        operation_mode=operation_mode
    )
    
    result = InventoryService.update_inventory_quantity(db, update_data, remark)
    op_label = "入库" if quantity_change > 0 else ("出库" if quantity_change < 0 else "调整")
    logger.info("库存%s: part_id=%d, 变动=%d, 备注=%s, 新库存=%d", op_label, part_id, quantity_change, remark, result.updatedQuantity)
    return {"success": True, "new_quantity": result.updatedQuantity, "message": "库存更新成功"}

@router.get("/inventory_history")
def get_inventory_history(
    part_id: int = Query(None, description="零件ID，不传则查询所有"),
    operation_type: str = Query(None, description="操作类型: in/out/adjust"),
    page: int = Query(1, description="页码", ge=1),
    page_size: int = Query(20, description="每页数量", ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取库存操作历史记录"""
    from app.crud.inventory_history import (
        get_inventory_history_by_part_id,
        get_all_inventory_history,
        get_inventory_history_count
    )
    
    offset = (page - 1) * page_size
    
    # 根据是否指定part_id选择不同的查询方法
    if part_id:
        history_list = get_inventory_history_by_part_id(db, part_id, limit=page_size, offset=offset)
        total_count = get_inventory_history_count(db, part_id)
    else:
        history_list = get_all_inventory_history(db, limit=page_size, offset=offset, operation_type=operation_type)
        total_count = get_inventory_history_count(db)
    
    # 构建响应数据
    result = [{
        "id": h.id,
        "part_id": h.part_id,
        "part_name": h.part.name if h.part else "Unknown",
        "operation_type": h.operation_type,
        "quantity_change": h.quantity_change,
        "quantity_before": h.quantity_before,
        "quantity_after": h.quantity_after,
        "remark": h.remark,
        "created_at": h.created_at.isoformat() if h.created_at else None
    } for h in history_list]
    
    return {
        "data": result,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": (total_count + page_size - 1) // page_size
        }
    }

@router.get("/get_parts_inventory")
def get_parts_inventory_list(
    page: int = Query(1, description="页码", ge=1),
    page_size: int = Query(100, description="每页数量", ge=1, le=500),
    sort_field: str = Query(None, description="排序字段"),
    sort_direction: str = Query("asc", description="排序方向", pattern="^(asc|desc|)$"),
    db: Session = Depends(get_db)
):
    """获取零件库存列表（支持分页和排序）"""
    if sort_field is None:
        sort_direction = None
    return InventoryService.get_parts_inventory_list(db, page, page_size, sort_field, sort_direction)

# ==================== 文件导入导出接口 ====================

@router.post("/import_order_excel_file")
async def import_order_excel_file(
    order_file: UploadFile = File(...),
    import_mode: str = Form("append"),
    db: Session = Depends(get_db),
    user=Depends(get_current_user_required)
):
    """导入Excel文件"""
    logger.info("开始导入Excel文件: %s, 模式: %s", order_file.filename, import_mode)
    try:
        content = await order_file.read()
        report = FileService.import_order_excel_file_direct(content, db, import_mode)
        logger.info("Excel导入完成: %s, 导入=%d, 跳过=%d, 错误=%d",
                     order_file.filename, report["imported"],
                     report["skipped_empty"] + report["skipped_no_name"] + report["skipped_no_quantity"] + report["skipped_bad_quantity"],
                     len(report["errors"]))
        return report
    except Exception as e:
        logger.error("Excel文件导入失败: %s, 错误: %s", order_file.filename, str(e))
        raise

@router.get("/get_part_by_id")
def get_part_by_id(part_id: int = Query(..., description="ID of the part to retrieve"), db: Session = Depends(get_db)):
    """根据ID获取零件详情"""
    return InventoryService.get_part_by_id(db, part_id)

@router.post("/update_part")
def update_part(
    part_id: int = Form(...),
    name: str = Form(...),
    manufacturer: str = Form(...),
    package: str = Form(...),
    price: str = Form(None),
    lc_number: str = Form(None),
    description: str = Form(None),
    other: str = Form(None),
    part_number: str = Form(None),
    part_type: str | None = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user_required)
):
    """更新零件信息"""
    result = InventoryService.update_part(db, part_id, name, manufacturer, package, price, lc_number, description, other, part_number)

    # 处理零件类型
    if part_type is not None and part_type.strip():
        from app.crud.type import get_type_by_name, create_part_type
        from app.models.part import Part
        existing_type = get_type_by_name(db, part_type.strip())
        if existing_type is None:
            existing_type = create_part_type(db, part_type.strip())
        part = db.query(Part).filter_by(id=part_id).first()
        if part:
            part.type_id = existing_type.id
            db.commit()

    logger.info("零件更新: part_id=%d, name=%s, manufacturer=%s, package=%s", part_id, name, manufacturer, package)
    return result

@router.post("/fix_missing_part_numbers")
def fix_missing_part_numbers(db: Session = Depends(get_db), user=Depends(get_current_user_required)):
    """批量修复无编号的零件"""
    return InventoryService.fix_missing_part_numbers(db)

@router.get("/search")
def search_in_inventory(search_key:str, db: Session = Depends(get_db)):
    """搜索零件"""
    return InventoryService.search(search_key, db)

@router.post("/advanced_search")
def advanced_search_in_inventory(
    filter_data: PartInventoryFilter,
    sort_field: str = Query(None, description="排序字段"),
    sort_direction: str = Query("asc", description="排序方向", pattern="^(asc|desc|)$"),
    db: Session = Depends(get_db)
):
    """高级筛选（支持多条件组合和排序）"""
    if sort_field is None:
        sort_direction = None
    return InventoryService.advanced_search(db, filter_data, sort_field, sort_direction)

# ==================== 基础数据接口 ====================

@router.get("/category_param_values")
def get_category_param_values(category_id: int = Query(...), db: Session = Depends(get_db)):
    """获取指定类别下零件的参数值分布（用于侧栏参数筛选）"""
    return InventoryService.get_category_param_values(db, category_id)

@router.get("/manufacturers")
def get_all_manufacturers(db: Session = Depends(get_db)):
    """获取所有制造商列表"""
    manufacturers = InventoryService.get_all_manufacturers(db)
    return [{"id": m.id, "name": m.name} for m in manufacturers]

@router.get("/packages")
def get_all_packages(db: Session = Depends(get_db)):
    """获取所有封装类型列表"""
    packages = InventoryService.get_all_packages(db)
    return [{"id": p.id, "name": p.package_type} for p in packages]

@router.get("/types")
def get_all_types(db: Session = Depends(get_db)):
    """获取所有零件类型列表"""
    types = InventoryService.get_all_types(db)
    return [{"id": t.id, "name": t.part_type} for t in types]
    

@router.delete("/delete_part")
def delete_part_with_id(part_id:int, db: Session = Depends(get_db), user=Depends(get_current_user_required)):
    """删除指定零件"""
    InventoryService.delete_part_with_id(part_id, db)
    logger.info("零件删除: part_id=%d", part_id)
    return {"message": f"Part with ID {part_id} deleted successfully"}

# ==================== 数据导出接口 ====================

def _get_inventory_data(db: Session) -> List:
    """获取所有库存数据的通用函数"""
    inventory_result = InventoryService.get_parts_inventory_list(db, page=1, page_size=10000)
    return inventory_result["data"]


def _add_type_dropdowns(workbook: Workbook, worksheet, db: Session, type_col: str, subtype_col: str, max_row: int = 200):
    """为类型/子类型列添加级联下拉选项
    - type_col: 类型列字母（如 'C'）
    - subtype_col: 子类型列字母（如 'D'）
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import get_column_letter

    # 查询所有类别和子类别
    categories = db.query(Category).order_by(Category.id).all()
    subcats = db.query(Subcategory).order_by(Subcategory.category_id, Subcategory.id).all()

    # 构建子类别映射：category_id -> [子类别名]
    subcat_map = {}
    for s in subcats:
        subcat_map.setdefault(s.category_id, []).append(s.name)

    if not categories:
        return

    # 创建隐藏的"下拉选项"工作表
    if "下拉选项" in workbook.sheetnames:
        options_ws = workbook["下拉选项"]
    else:
        options_ws = workbook.create_sheet("下拉选项")
    options_ws.sheet_state = "hidden"

    # 第1列：所有类别名称（从第2行开始）
    for i, cat in enumerate(categories, start=2):
        options_ws.cell(row=i, column=1, value=cat.name)

    # 每个类别一个子类别列表（列2开始，列头为类别名）
    for col_idx, cat in enumerate(categories, start=2):
        options_ws.cell(row=1, column=col_idx, value=cat.name)
        for row_idx, sub_name in enumerate(subcat_map.get(cat.id, []), start=2):
            options_ws.cell(row=row_idx, column=col_idx, value=sub_name)

    # 定义命名区域：类型列表
    last_cat_row = len(categories) + 1
    workbook.defined_names.add(
        DefinedName("类型列表", attr_text=f"下拉选项!$A$2:$A${last_cat_row}")
    )

    # 每个类别定义一个命名区域：类别名 → 子类别列表
    for col_idx, cat in enumerate(categories, start=2):
        sub_count = len(subcat_map.get(cat.id, []))
        if sub_count == 0:
            continue
        col_letter = get_column_letter(col_idx)
        last_sub_row = sub_count + 1
        # 命名区域名用类别的ASCII安全形式（避免中文命名兼容性问题）
        safe_name = f"SUBTYPE_{cat.id}"
        workbook.defined_names.add(
            DefinedName(safe_name, attr_text=f"下拉选项!${col_letter}$2:${col_letter}${last_sub_row}")
        )

    # 类型列数据验证（直接引用类别列表）
    type_dv = DataValidation(
        type="list",
        formula1="=类型列表",
        allow_blank=True,
        showDropDown=False
    )
    worksheet.add_data_validation(type_dv)
    type_dv.add(f"{type_col}2:{type_col}{max_row}")

    # 子类型列：根据左侧类型列的值动态显示对应子类别
    # 使用 INDIRECT 引用命名区域，命名区域名需要与类型值匹配
    # 由于命名区域使用 SUBTYPE_{id} 形式，这里通过辅助公式映射
    # 在隐藏工作表中添加映射列：类别名 -> 命名区域名
    map_col = len(categories) + 2  # 映射列
    options_ws.cell(row=1, column=map_col, value="类型名")
    options_ws.cell(row=1, column=map_col + 1, value="区域名")
    for i, cat in enumerate(categories, start=2):
        options_ws.cell(row=i, column=map_col, value=cat.name)
        options_ws.cell(row=i, column=map_col + 1, value=f"SUBTYPE_{cat.id}")

    # 映射命名区域
    map_last_row = len(categories) + 1
    map_col_letter = get_column_letter(map_col)
    map_range_letter = get_column_letter(map_col + 1)
    workbook.defined_names.add(
        DefinedName("类型到区域映射", attr_text=f"下拉选项!${map_col_letter}$2:${map_range_letter}${map_last_row}")
    )

    # 子类型列使用 VLOOKUP 找到对应命名区域名，再用 INDIRECT 引用
    subtype_dv = DataValidation(
        type="list",
        formula1=f"=INDIRECT(VLOOKUP({type_col}2,类型到区域映射,2,FALSE))",
        allow_blank=True,
        showDropDown=False
    )
    worksheet.add_data_validation(subtype_dv)
    subtype_dv.add(f"{subtype_col}2:{subtype_col}{max_row}")


@router.get("/export_template_excel")
def export_import_template_excel(db: Session = Depends(get_db)):
    """导出导入模板为Excel格式"""
    # 创建工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "导入模板"
    
    # 写入表头
    headers = ['Name', 'Manufacturer', 'Type', 'Subtype', 'Package', 'Quantity', 'LC Number', 'Price', 'Description']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)

    # 添加类型/子类型级联下拉（Type列=C, Subtype列=D）
    _add_type_dropdowns(workbook, worksheet, db, "C", "D", max_row=200)
    
    # 写入示例行
    worksheet.cell(row=2, column=1, value='NE555')
    worksheet.cell(row=2, column=2, value='TI')
    worksheet.cell(row=2, column=3, value='IC')
    worksheet.cell(row=2, column=4, value='')
    worksheet.cell(row=2, column=5, value='DIP-8')
    worksheet.cell(row=2, column=6, value=10)
    worksheet.cell(row=2, column=7, value='C12345')
    worksheet.cell(row=2, column=8, value=0.5)
    worksheet.cell(row=2, column=9, value='Timer IC')
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    headers = {'Content-Disposition': 'attachment; filename="inventory_template.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@router.get("/export_excel")
def export_inventory_excel(db: Session = Depends(get_db)):
    """导出库存数据为Excel格式"""
    import json as _json
    inventory_data = _get_inventory_data(db)
    
    # 收集所有参数字段
    all_param_fields = set()
    for item in inventory_data:
        if item.other:
            try:
                data = _json.loads(item.other)
                if isinstance(data, dict):
                    if 'fields' in data and 'values' in data:
                        # 新格式
                        all_param_fields.update(data.get('values', {}).keys())
                    else:
                        # 旧格式
                        all_param_fields.update(data.keys())
            except:
                pass
    
    param_fields = sorted(list(all_param_fields))
    
    # 创建工作簿和工作表
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "库存清单"
    
    # 写入表头
    headers = ['编号', 'Name', 'Manufacturer', '类型', '子类型', 'Package', 'Quantity', 'LC Number', 'Price', 'Description']
    headers.extend(param_fields)
    
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
        worksheet.cell(row=1, column=col).font = worksheet.cell(row=1, column=col).font.copy(bold=True)
    
    # 写入数据
    for row_idx, item in enumerate(inventory_data, 2):
        # 类型/子类型分开
        cat_str = ''
        sub_str = ''
        if item.category_name:
            cat_str = item.category_name
            sub_str = item.subcategory_name or ''
        else:
            cat_str = item.part_type or ''
        
        worksheet.cell(row=row_idx, column=1, value=item.part_number or '')
        worksheet.cell(row=row_idx, column=2, value=item.name)
        worksheet.cell(row=row_idx, column=3, value=item.manufacturer)
        worksheet.cell(row=row_idx, column=4, value=cat_str)
        worksheet.cell(row=row_idx, column=5, value=sub_str)
        worksheet.cell(row=row_idx, column=6, value=item.package)
        worksheet.cell(row=row_idx, column=7, value=item.quantity)
        worksheet.cell(row=row_idx, column=8, value=item.lc_number if item.lc_number else '')
        worksheet.cell(row=row_idx, column=9, value=item.price_display if item.price_display else (item.price if item.price else ''))
        worksheet.cell(row=row_idx, column=10, value=item.description if item.description else '')
        
        # 解析参数
        if item.other:
            try:
                data = _json.loads(item.other)
                if isinstance(data, dict):
                    params = {}
                    if 'fields' in data and 'values' in data:
                        # 新格式
                        params = data.get('values', {})
                    else:
                        # 旧格式
                        params = data
                    
                    for param_name, param_value in params.items():
                        if param_name in param_fields:
                            col_idx = headers.index(param_name) + 1
                            worksheet.cell(row=row_idx, column=col_idx, value=str(param_value))
            except:
                pass
    
    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 添加类型/子类型级联下拉（类型列=D, 子类型列=E）
    max_data_rows = len(inventory_data) + 1
    _add_type_dropdowns(workbook, worksheet, db, "D", "E", max_row=max_data_rows)

    # 保存到内存
    excel_output = io.BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)
    
    headers = {'Content-Disposition': 'attachment; filename="inventory_export.xlsx"'}
    return StreamingResponse(
        excel_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )