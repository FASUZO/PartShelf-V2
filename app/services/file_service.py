import io
import json
import logging
import re
import zipfile
import xml.etree.ElementTree as ET
from fastapi import HTTPException, status
from sqlalchemy.orm import Session
from app.crud.file_templates import create_file_template, get_available_file_templates, get_template_by_id
from app.models.file_template import FileTemplate
from app.models.part import Part
from app.models.inventory import Inventory
from app.models.inventory_history import InventoryHistory
from app.models.manufacturer import Manufacturer
from app.models.package import Package as PackageModel
from app.models.type import Type
from app.models.config import Category, Subcategory
from app.schemas.file_template import FileTemplateAdd, FileTemplateGet
from app.schemas.inventory import PartToInventoryAdd
from app.services.inventory_service import InventoryService

logger = logging.getLogger("partshelf.import")

class FileService:
    @staticmethod
    def match_category_by_type(db: Session, part_type: str):
        """根据类型名称匹配类别ID，支持'类别/子类别'格式和模糊匹配"""
        if not part_type:
            return None, None
        
        # 支持"类别/子类别"格式
        if '/' in part_type:
            parts = part_type.split('/', 1)
            cat_name = parts[0].strip()
            sub_name = parts[1].strip() if len(parts) > 1 else ''
            
            # 匹配类别
            category = db.query(Category).filter(Category.name == cat_name).first()
            if not category:
                # 模糊匹配
                for cat in db.query(Category).all():
                    if cat.name in cat_name or cat_name in cat.name:
                        category = cat
                        break
            
            if category and sub_name:
                # 匹配子类别
                subcategory = db.query(Subcategory).filter(
                    Subcategory.category_id == category.id,
                    Subcategory.name == sub_name
                ).first()
                if not subcategory:
                    # 模糊匹配
                    for sub in db.query(Subcategory).filter(Subcategory.category_id == category.id).all():
                        if sub.name in sub_name or sub_name in sub.name:
                            subcategory = sub
                            break
                
                if subcategory:
                    return category.id, subcategory.id
            
            if category:
                return category.id, None
        
        # 精确匹配类别名称
        category = db.query(Category).filter(Category.name == part_type).first()
        if category:
            return category.id, None
        
        # 模糊匹配类别名称（包含关系）
        categories = db.query(Category).all()
        for cat in categories:
            if cat.name in part_type or part_type in cat.name:
                return cat.id, None
        
        # 尝试匹配子类别
        subcategory = db.query(Subcategory).filter(Subcategory.name == part_type).first()
        if subcategory:
            return subcategory.category_id, subcategory.id
        
        # 模糊匹配子类别
        subcategories = db.query(Subcategory).all()
        for sub in subcategories:
            if sub.name in part_type or part_type in sub.name:
                return sub.category_id, sub.id
        
        return None, None

    def add_file_template(db: Session, template: FileTemplateAdd):
        return create_file_template(db ,FileTemplate(
            template_type = template.template_type,
            template_name = template.template_name,
            manufacturer_column = template.manufacturer_column,
            part_name_column =  template.part_name_column,
            package_column = template.package_column,
            description_column = template.description_column,
            quantity_column =template.quantity_column
        ))
    
    def get_available_file_templates(db: Session):
        file_templates = get_available_file_templates(db)
        
        return [
        FileTemplateGet(
            id = template.id,
            template_type = template.template_type,
            template_name = template.template_name,
            
            manufacturer_column = template.manufacturer_column,
            part_name_column = template.part_name_column,
            package_column = template.package_column,
            description_column = template.description_column,
            quantity_column = template.quantity_column,
        )
        for template in file_templates
        ] 
    
    def extract_info(description):
        if description is None or description == "" or str(description).strip() == "":
            return None, None
        
        value_match = re.search(r'\b\d+(?:\.\d+)?\s?(kΩ|Ω|MΩ|nF|uF|pF|F|H|mH|µH)\b', description, re.IGNORECASE)
        value = value_match.group(0) if value_match else ""

        type_match = re.search(r'(Thick Film Resistor|Chip Resistor|Capacitor|Multilayer Ceramic Capacitor|Operational Amplifier)', description, re.IGNORECASE)
        part_type = type_match.group(0) if type_match else ""
        
        return value, part_type

    def parse_excel_file(file_path):
        """解析Excel文件"""
        try:
            # 优先使用openpyxl直接解析（兼容性更好）
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            wb.close()
            return data
        except Exception:
            try:
                import pandas as pd
                df = pd.read_excel(file_path, header=None)
                data = df.values.tolist()
                return data
            except Exception:
                import zipfile
                import xml.etree.ElementTree as ET
            
            data = []
            
            # 打开Excel文件（实际上是zip文件）
            with zipfile.ZipFile(file_path, 'r') as zf:
                # 读取sharedStrings.xml（共享字符串表）
                shared_strings = {}
                if 'xl/sharedStrings.xml' in zf.namelist():
                    shared_strings_content = zf.read('xl/sharedStrings.xml').decode('utf-8')
                    root = ET.fromstring(shared_strings_content)
                    for i, si in enumerate(root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si')):
                        t = si.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                        if t is not None and t.text:
                            shared_strings[i] = t.text
                
                # 读取第一个工作表
                sheet_files = [f for f in zf.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
                if not sheet_files:
                    return data
                
                # 读取第一个工作表
                sheet_content = zf.read(sheet_files[0]).decode('utf-8')
                root = ET.fromstring(sheet_content)
                
                # 解析工作表数据
                sheet_data = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetData')
                if sheet_data is None:
                    return data
                
                # 解析每一行
                for row_elem in sheet_data.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                    row_data = []
                    
                    # 解析每个单元格
                    for cell_elem in row_elem.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                        cell_value = ""
                        v_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        
                        if v_elem is not None and v_elem.text:
                            cell_type = cell_elem.get('t')
                            if cell_type == 's':  # 共享字符串
                                idx = int(v_elem.text)
                                cell_value = shared_strings.get(idx, "")
                            else:  # 直接值
                                cell_value = v_elem.text
                        
                        row_data.append(cell_value)
                    
                    data.append(row_data)
            
            return data

    def import_order_excel_file(file_content: bytes, template_id: int, db: Session):
        try:
            # 使用zipfile和xml解析Excel文件，避免外部库依赖问题
            import tempfile
            import os
            
            # 创建临时文件
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(file_content)
                tmp_file_path = tmp_file.name
            
            try:
                # 解析Excel文件
                data = FileService.parse_excel_file(tmp_file_path)
                
                # 获取文件模板
                file_template = get_template_by_id(db, template_id)
                if not file_template:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"Template with id={template_id} not found to parse file"
                    )
                
                # 获取列名（第一行）
                if not data or len(data) < 2:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel文件为空或格式不正确"
                    )
                
                header_row = data[0]
                
                # 找到对应的列索引
                manufacturer_col_idx = None
                part_name_col_idx = None
                package_col_idx = None
                description_col_idx = None
                quantity_col_idx = None
                
                for idx, col_name in enumerate(header_row):
                    if col_name == file_template.manufacturer_column:
                        manufacturer_col_idx = idx
                    elif col_name == file_template.part_name_column:
                        part_name_col_idx = idx
                    elif col_name == file_template.package_column:
                        package_col_idx = idx
                    elif col_name == file_template.description_column:
                        description_col_idx = idx
                    elif col_name == file_template.quantity_column:
                        quantity_col_idx = idx
                
                # 检查是否找到了所有必需的列
                if not all([manufacturer_col_idx is not None, part_name_col_idx is not None, 
                           package_col_idx is not None, quantity_col_idx is not None]):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel文件中缺少必需的列"
                    )
                
                # 遍历数据行（从第二行开始，假设第一行是列名）
                for row_idx, row in enumerate(data[1:], start=2):
                    # 跳过空行
                    if not row or len(row) <= max(manufacturer_col_idx, part_name_col_idx, package_col_idx, quantity_col_idx):
                        continue
                    
                    # 获取行数据
                    manufacturer = row[manufacturer_col_idx] if manufacturer_col_idx < len(row) else None
                    part_name = row[part_name_col_idx] if part_name_col_idx < len(row) else None
                    package = row[package_col_idx] if package_col_idx < len(row) else None
                    quantity = row[quantity_col_idx] if quantity_col_idx < len(row) else None
                    description = str(row[description_col_idx]).strip() if description_col_idx is not None and description_col_idx < len(row) and row[description_col_idx] else ""
                    
                    # 跳过空行 - 根据要求只需要型号列和数量列有数据
                    if not part_name or not quantity:
                        continue
                    
                    # 清理数据
                    manufacturer = str(manufacturer).strip()
                    part_name = str(part_name).strip()
                    package = str(package).strip()
                    quantity = int(float(quantity)) if isinstance(quantity, (int, float)) else int(quantity)
                    
                    # 提取信息
                    value, part_type = FileService.extract_info(description)
                    
                    # 创建零件数据
                    part_to_add = PartToInventoryAdd(
                        name=part_name,
                        manufacturer=manufacturer,
                        package=package,
                        quantity=quantity,
                        description=description,
                        part_type=part_type
                    )
                    
                    # 添加到库存
                    InventoryService.add_part_to_inventory(db, part_to_add, record_history=False)
                
                return {"message": "Excel文件导入成功"}
                
            finally:
                # 删除临时文件
                os.unlink(tmp_file_path)
            
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Excel文件导入失败: {str(e)}"
            )
    
    def import_order_excel_file_direct(file_content: bytes, db: Session, import_mode: str = "append"):
        """直接导入Excel文件，使用预设的列映射

        Args:
            file_content: Excel文件内容
            db: 数据库会话
            import_mode: 导入模式，"append"（追加）或 "overwrite"（覆盖）

        Returns:
            dict: 导入报告
        """
        report = {
            "total_rows": 0,
            "imported": 0,
            "updated": 0,
            "skipped_empty": 0,
            "skipped_no_name": 0,
            "skipped_no_quantity": 0,
            "skipped_bad_quantity": 0,
            "errors": [],
            "details": [],
            "columns_detected": [],
            "columns_mapped": {},
            "mode": import_mode,
        }

        try:
            import tempfile
            import os

            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(file_content)
                tmp_file_path = tmp_file.name

            try:
                data = FileService.parse_excel_file(tmp_file_path)
                
                # 预设的列名映射（支持中英文，不区分大小写）
                column_mapping = {
                    'name': ['name', '型号', '商品型号', '名称', 'product name', 'part name'],
                    'manufacturer': ['manufacturer', '制造商', '厂家', '厂商', '品牌', 'brand'],
                    'package': ['package', '封装', '封装类型', 'package type', '封装格式', 'footprint'],
                    'quantity': ['quantity', '数量', 'qty', '库存', '型号发货数量', 'count', 'amount'],
                    'description': ['description', '描述', '说明', 'desc', 'details'],
                    'part_type': ['part_type', '类型', '商品类型', '零件类型', 'type', 'category'],
                    'subcategory': ['subcategory', '子类型', '子类别', 'sub_type', 'subtype'],
                    'price': ['price', '单价', '单价（人民币含税）', 'unit price', 'cost'],
                    'lc_number': ['lc_number', 'LC编号', '商品编号', 'lc number', 'part number'],
                    'part_number': ['part_number', '编号', '位号', '零件编号'],
                    'other': ['other', '其他', '其他信息', 'remarks', 'note']
                }

                # 获取列名，处理第一行可能是空行的情况
                if not data or len(data) < 2:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel文件为空或格式不正确"
                    )

                # 直接使用第一行作为表头（导出的文件第一行就是表头）
                header_row = data[0]
                data_start_row = 0

                # 如果第一行没有有效数据，尝试第二行
                if not any(cell and str(cell).strip() for cell in header_row):
                    if len(data) > 1:
                        header_row = data[1]
                        data_start_row = 1

                if not any(cell and str(cell).strip() for cell in header_row):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel文件中没有找到有效的列名行"
                    )
                
                # 自动检测列映射
                col_mapping = {}
                for field, possible_names in column_mapping.items():
                    for idx, col_name in enumerate(header_row):
                        if col_name and any(name.lower() in str(col_name).lower() for name in possible_names):
                            col_mapping[field] = idx
                            break

                # 已映射的列索引（用于识别额外参数列）
                mapped_col_indices = set(col_mapping.values())

                # 记录检测到的列
                for idx, col_name in enumerate(header_row):
                    if col_name and str(col_name).strip():
                        report["columns_detected"].append(str(col_name).strip())
                for field, idx in col_mapping.items():
                    if idx < len(header_row) and header_row[idx]:
                        report["columns_mapped"][field] = str(header_row[idx]).strip()

                # 检查是否找到了所有必需的列
                required_fields = ['name', 'manufacturer', 'package', 'quantity']
                missing_fields = [f for f in required_fields if f not in col_mapping]
                if missing_fields:
                    # 提供更详细的错误信息，包括实际检测到的列名
                    detected_columns = []
                    for idx, col_name in enumerate(header_row):
                        if col_name and col_name.strip():
                            detected_columns.append(f"'{col_name}'")
                    
                    if detected_columns:
                        detail_msg = f"Excel文件中缺少必需的列: {', '.join(missing_fields)}。检测到的列名: {', '.join(detected_columns)}"
                    else:
                        detail_msg = f"Excel文件中缺少必需的列: {', '.join(missing_fields)}。第一行似乎是空的，请确保Excel文件包含列名。"
                    
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=detail_msg
                    )

                # 覆盖模式：验证通过后再清空数据（防止无效文件导致数据丢失）
                if import_mode == "overwrite":
                    from app.models.inventory_history import InventoryHistory
                    # 先删除历史记录（外键约束）
                    db.query(InventoryHistory).delete()
                    # 再删除库存和零件
                    db.query(Inventory).delete()
                    db.query(Part).delete()
                    db.commit()
                    report["overwrite_cleared"] = True
                    logger.info("覆盖模式：已清空库存和历史记录")

                # 遍历数据行（从列名行的下一行开始）
                for row_idx, row in enumerate(data[data_start_row + 1:], start=data_start_row + 2):
                    report["total_rows"] += 1

                    # 跳过空行
                    if not row or len(row) <= max(col_mapping.values()):
                        report["skipped_empty"] += 1
                        continue

                    # 获取行数据
                    row_data = {}
                    extra_params = {}
                    for field, col_idx in col_mapping.items():
                        if col_idx < len(row):
                            value = row[col_idx]
                            row_data[field] = str(value).strip() if value is not None else ""
                        else:
                            row_data[field] = ""

                    # 收集额外的参数列
                    for idx, cell in enumerate(row):
                        if idx not in mapped_col_indices and idx < len(header_row):
                            col_name = header_row[idx]
                            if col_name and str(col_name).strip():
                                value = str(cell).strip() if cell is not None else ""
                                if value:
                                    extra_params[str(col_name).strip()] = value

                    # 如果有额外参数，合并到 other 字段
                    if extra_params:
                        import json as _json
                        existing_other = row_data.get('other', '')
                        if existing_other:
                            try:
                                other_data = _json.loads(existing_other)
                                if isinstance(other_data, dict) and 'fields' in other_data:
                                    other_data.get('values', {}).update(extra_params)
                                else:
                                    other_data.update(extra_params)
                            except:
                                other_data = extra_params
                        else:
                            other_data = {
                                "fields": list(extra_params.keys()),
                                "values": extra_params,
                                "units": {}
                            }
                        row_data['other'] = _json.dumps(other_data, ensure_ascii=False)

                    # 检查是否有数据
                    has_data = any(row_data.get(field) for field in required_fields)
                    if not has_data:
                        report["skipped_empty"] += 1
                        continue

                    if not row_data.get('name'):
                        report["skipped_no_name"] += 1
                        report["details"].append({"row": row_idx, "status": "skipped", "reason": "型号为空"})
                        continue

                    if not row_data.get('quantity'):
                        report["skipped_no_quantity"] += 1
                        report["details"].append({"row": row_idx, "status": "skipped", "reason": "数量为空", "name": row_data.get('name', '')})
                        continue

                    try:
                        quantity = int(float(row_data['quantity']))
                    except ValueError:
                        report["skipped_bad_quantity"] += 1
                        report["details"].append({"row": row_idx, "status": "skipped", "reason": "数量格式错误: " + row_data['quantity'], "name": row_data.get('name', '')})
                        continue

                    description = row_data.get('description', '')
                    part_type = row_data.get('part_type', '')
                    if not part_type:
                        value, part_type = FileService.extract_info(description)

                    sub_type = row_data.get('subcategory', '')
                    if sub_type and part_type and '/' not in part_type:
                        part_type = f"{part_type}/{sub_type}"

                    category_id, subcategory_id = FileService.match_category_by_type(db, part_type)
                    if category_id:
                        logger.info(f"行{row_idx}: 类型'{part_type}'匹配到类别ID={category_id}, 子类别ID={subcategory_id}")

                    try:
                        part_to_add = PartToInventoryAdd(
                            name=row_data['name'],
                            manufacturer=row_data['manufacturer'],
                            package=row_data['package'],
                            quantity=quantity,
                            description=description,
                            part_type=part_type,
                            lc_number=row_data.get('lc_number', None),
                            price=InventoryService._parse_price(row_data.get('price')),
                            other=row_data.get('other', None),
                            category_id=category_id,
                            subcategory_id=subcategory_id
                        )
                        InventoryService.add_part_to_inventory(db, part_to_add, record_history=False)
                        report["imported"] += 1
                        report["details"].append({
                            "row": row_idx,
                            "status": "ok",
                            "name": row_data['name'],
                            "quantity": quantity,
                            "category": part_type or "-",
                        })
                    except Exception as e:
                        report["errors"].append({"row": row_idx, "name": row_data.get('name', ''), "error": str(e)})
                        report["details"].append({"row": row_idx, "status": "error", "name": row_data.get('name', ''), "reason": str(e)})

                # 导入后数据库验证：检查实际入库数据是否完整
                for detail in report["details"]:
                    if detail["status"] != "ok":
                        detail["verify_status"] = "跳过验证"
                        detail["verify_issues"] = ""
                        continue

                    name = detail.get("name", "")
                    issues = []

                    # 查找刚导入的零件（按名称+数量匹配最近的记录）
                    part = db.query(Part).filter(Part.name == name).order_by(Part.id.desc()).first()
                    if not part:
                        detail["verify_status"] = "未找到"
                        detail["verify_issues"] = "数据库中未找到该零件"
                        continue

                    # 检查库存记录
                    inv = db.query(Inventory).filter(Inventory.part_id == part.id).first()
                    if not inv:
                        issues.append("缺少库存记录")
                    elif inv.quantity_available != detail.get("quantity", 0):
                        issues.append("数量不匹配: 期望%d, 实际%d" % (detail.get("quantity", 0), inv.quantity_available))

                    # 检查制造商
                    if not part.manufacturer_id:
                        mfr = db.query(Manufacturer).filter(Manufacturer.name == detail.get("manufacturer", "")).first()
                        if not mfr:
                            issues.append("制造商未关联")
                    else:
                        mfr = db.query(Manufacturer).filter(Manufacturer.id == part.manufacturer_id).first()
                        detail["manufacturer_verified"] = mfr.name if mfr else "未知"

                    # 检查封装
                    if not part.package_id:
                        issues.append("封装未关联")

                    # 检查类别
                    if not part.category_id:
                        issues.append("未分配类别")
                    else:
                        cat = db.query(Category).filter(Category.id == part.category_id).first()
                        detail["category_verified"] = cat.name if cat else "未知ID:%d" % part.category_id
                        if not cat:
                            issues.append("类别ID=%d不存在" % part.category_id)

                        if part.subcategory_id:
                            subcat = db.query(Subcategory).filter(Subcategory.id == part.subcategory_id).first()
                            if subcat:
                                detail["subcategory_verified"] = subcat.name
                                # 验证子类别是否属于该类别
                                if cat and subcat.category_id != cat.id:
                                    issues.append("子类别'%s'不属于类别'%s'" % (subcat.name, cat.name))
                            else:
                                detail["subcategory_verified"] = "未知ID:%d" % part.subcategory_id
                                issues.append("子类别ID=%d不存在" % part.subcategory_id)
                        else:
                            # 没有子类别 — 标记为问题
                            expected_type = detail.get("category", "")
                            if "/" in expected_type:
                                expected_sub = expected_type.split("/", 1)[1].strip()
                                issues.append("子类别'%s'未匹配到" % expected_sub)
                            else:
                                issues.append("未分配子类别")

                    # 检查零件编号
                    if not part.part_number:
                        issues.append("未生成零件编号")

                    detail["verify_status"] = "通过" if not issues else "有问题"
                    detail["verify_issues"] = "; ".join(issues)

                # 统计验证结果
                report["verify_passed"] = sum(1 for d in report["details"] if d.get("verify_status") == "通过")
                report["verify_issues"] = sum(1 for d in report["details"] if d.get("verify_status") == "有问题")

                # 发送MQTT通知
                try:
                    from app.services.mqtt_service import publish_event
                    import json
                    payload = json.dumps({
                        "type": "excel",
                        "rows": row_idx - data_start_row - 1,
                        "mode": import_mode
                    }, ensure_ascii=False)
                    publish_event("inventory.import", payload)
                except Exception:
                    pass

                return report

            finally:
                os.unlink(tmp_file_path)
            
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Excel文件导入失败: {str(e)}"
            )
        
